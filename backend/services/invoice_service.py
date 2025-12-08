"""Invoice management and generation service."""

import os
from typing import List, Optional, Dict, Tuple, Union
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
import pytz
from rapidfuzz import fuzz

from database.models import Invoice, InvoiceItem, Product, Customer
from utils.text_utils import normalize_vietnamese, normalize_phone
from config import Config
from pathlib import Path

# UTC+7 timezone for Vietnam
VN_TZ = pytz.timezone('Asia/Ho_Chi_Minh')
ALLOWED_STATUSES = {'pending', 'paid', 'cancelled', 'processing'}

def get_vn_time():
    """Get current time in Vietnam timezone (UTC+7)."""
    return datetime.now(VN_TZ)


def number_to_vietnamese_text(number: float) -> str:
    """
    Convert a number to Vietnamese text representation.
    
    Args:
        number: The number to convert (e.g., 218028000)
    
    Returns:
        Vietnamese text representation (e.g., "Hai trăm mười tám triệu hai mươi tám nghìn đồng")
    
    Examples:
        218028000 -> "Hai trăm mười tám triệu hai mươi tám nghìn đồng"
        1872505 -> "Một triệu tám trăm bảy mươi hai nghìn năm trăm linh năm đồng"
        100001 -> "Một trăm nghìn không trăm linh một đồng"
    """
    if number == 0:
        return "Không đồng"
    
    # Handle negative numbers
    if number < 0:
        return "Âm " + number_to_vietnamese_text(-number)
    
    # Round to integer
    number = int(round(number))
    
    # Number words
    ones = ["", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
    
    def read_two_digits(n, is_last_two_of_group):
        """
        Read the last two digits (tens and ones).
        
        Args:
            n: Number from 0-99
            is_last_two_of_group: True if these are the last 2 digits of a 3-digit group with hundreds > 0
        
        Returns:
            Vietnamese text for the two digits
        """
        result = []
        ten = n // 10
        one = n % 10
        
        if ten == 0:
            if one > 0:
                if is_last_two_of_group:
                    result.append("linh")
                result.append(ones[one])
        elif ten == 1:
            result.append("mười")
            if one == 5:
                result.append("lăm")
            elif one > 0:
                result.append(ones[one])
        else:  # ten >= 2
            result.append(ones[ten])
            result.append("mươi")
            if one == 1:
                result.append("mốt")
            elif one == 5:
                result.append("lăm")
            elif one > 0:
                result.append(ones[one])
        
        return " ".join(result)
    
    def read_three_digits(n):
        """
        Read a group of 3 digits (000-999).
        
        Args:
            n: Number from 0-999
        
        Returns:
            Vietnamese text for the three digits
        """
        if n == 0:
            return ""
        
        hundred = n // 100
        remainder = n % 100
        
        result = []
        
        if hundred > 0:
            result.append(ones[hundred])
            result.append("trăm")
            if remainder > 0:
                result.append(read_two_digits(remainder, is_last_two_of_group=True))
        else:
            # No hundreds, just read tens and ones normally
            result.append(read_two_digits(remainder, is_last_two_of_group=False))
        
        return " ".join(result)
    
    # Break number into groups of 3 digits
    groups = []
    group_names = ["", "nghìn", "triệu", "tỷ"]
    
    # Extract groups from right to left
    temp = number
    while temp > 0:
        groups.append(temp % 1000)
        temp //= 1000
    
    # Pad to have at least 1 group
    if not groups:
        groups = [0]
    
    # Reverse to process from left to right (largest to smallest)
    groups.reverse()
    
    # Read groups
    result = []
    num_groups = len(groups)
    
    for i, group_value in enumerate(groups):
        if group_value > 0:
            group_text = read_three_digits(group_value)
            result.append(group_text)
            
            # Add group name (nghìn, triệu, tỷ)
            group_index = num_groups - i - 1
            if group_index > 0 and group_index < len(group_names):
                result.append(group_names[group_index])
    
    text = " ".join(result)
    
    # Capitalize first letter
    if text:
        text = text[0].upper() + text[1:]
    
    return text + " đồng"
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Register fonts for Vietnamese support.
# On Windows the Linux font path does not exist, so we keep the files inside the repo and also try Windows system fonts.
FONT_DIR = Path(__file__).resolve().parent / "fonts"
DEJAVU_REGULAR_URL = "https://dejavu-fonts.github.io/dejavu-fonts/ttf/DejaVuSans.ttf"
DEJAVU_BOLD_URL = "https://dejavu-fonts.github.io/dejavu-fonts/ttf/DejaVuSans-Bold.ttf"


def _ensure_font(path: Path, url: str):
    """Download a font if it is missing so Vietnamese text renders correctly."""
    if path.exists():
        return
    try:
        FONT_DIR.mkdir(parents=True, exist_ok=True)
        import requests
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        path.write_bytes(response.content)
    except Exception as e:
        print(f"Warning: Could not download font {path.name}: {e}")


def _is_valid_ttf(path: Path) -> bool:
    """Quick check to avoid HTML/error files when download fails."""
    try:
        if not path.exists() or path.stat().st_size < 1024:
            return False
        with path.open("rb") as f:
            header = f.read(4)
            return header in {b"\x00\x01\x00\x00", b"OTTO"}
    except Exception:
        return False


regular_font_path = FONT_DIR / "DejaVuSans.ttf"
bold_font_path = FONT_DIR / "DejaVuSans-Bold.ttf"
_ensure_font(regular_font_path, DEJAVU_REGULAR_URL)
_ensure_font(bold_font_path, DEJAVU_BOLD_URL)

# Windows system fonts as backup (have Vietnamese glyphs)
windows_regular = Path("C:/Windows/Fonts/arial.ttf")
windows_bold = Path("C:/Windows/Fonts/arialbd.ttf")

# Try repo fonts first, then Windows fonts, otherwise fall back to Helvetica
FONT_NORMAL = 'Helvetica'
FONT_BOLD = 'Helvetica-Bold'
font_candidates = [
    ("DejaVuSans", regular_font_path, bold_font_path),
    ("Arial", windows_regular, windows_bold),
]

for font_name, normal_path, bold_path in font_candidates:
    if not (_is_valid_ttf(normal_path) and _is_valid_ttf(bold_path)):
        continue
    try:
        pdfmetrics.registerFont(TTFont(font_name, str(normal_path)))
        pdfmetrics.registerFont(TTFont(f"{font_name}-Bold", str(bold_path)))
        pdfmetrics.registerFontFamily(
            font_name,
            normal=font_name,
            bold=f"{font_name}-Bold",
            italic=font_name,  # Fallback
            boldItalic=f"{font_name}-Bold"  # Fallback
        )
        FONT_NORMAL = font_name
        FONT_BOLD = f"{font_name}-Bold"
        break
    except Exception as e:
        print(f"Warning: Could not load {font_name} fonts from {normal_path}: {e}")


class InvoiceService:
    """Service for invoice management and generation."""

    def __init__(self, db_session: Session, output_dir: str = "data/invoices"):
        """
        Initialize invoice service.

        Args:
            db_session: SQLAlchemy database session
            output_dir: Directory to save generated invoices
        """
        self.db = db_session
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def _normalize_page_params(self, limit: int, offset: int, max_limit: int = max(Config.MAX_PAGE_SIZE, 100)) -> tuple[int, int]:
        """Clamp pagination inputs to safe bounds."""
        safe_limit = max(1, min(limit or 0, max_limit))
        safe_offset = max(0, offset or 0)
        return safe_limit, safe_offset

    def generate_invoice_number(self) -> str:
        """
        Generate a unique invoice number using Vietnam timezone (UTC+7).

        Returns:
            Invoice number in format INV-YYYYMMDD-XXXX
        """
        today = get_vn_time().strftime("%Y%m%d")
        prefix = f"INV-{today}"

        # Count invoices created today
        count = self.db.query(Invoice).filter(
            Invoice.invoice_number.like(f"{prefix}%")
        ).count()

        return f"{prefix}-{count + 1:04d}"

    def create_invoice(
        self,
        items: List[Dict],
        customer_id: Optional[int] = None,
        customer_name: Optional[str] = None,
        customer_phone: Optional[str] = None,
        customer_address: Optional[str] = None,
        discount: float = 0,
        tax: float = 0,
        payment_method: Optional[str] = None,
        notes: Optional[str] = None,
        status: str = 'pending'
    ) -> Invoice:
        """
        Create a new invoice.

        Args:
            items: List of dicts with 'product_id' and 'quantity'
            customer_id: Customer ID (if registered customer)
            customer_name: Customer name (for one-time customers)
            customer_phone: Customer phone
            customer_address: Customer address
            discount: Discount amount
            tax: Tax amount
            payment_method: Payment method
            notes: Additional notes
            status: Invoice status (pending, paid, cancelled)

        Returns:
            Created Invoice object
        """
        if status not in ALLOWED_STATUSES:
            raise ValueError("Trạng thái hóa đơn không hợp lệ.")

        # Allow empty items only when saving a draft invoice
        if status != 'processing' and len(items) == 0:
            raise ValueError("Hóa đơn cần ít nhất 1 sản phẩm (trừ khi lưu ở trạng thái chờ xử lý).")

        # Draft invoices shouldn't carry over discount/tax to avoid negative totals
        if status == 'processing' and len(items) == 0:
            discount = 0
            tax = 0

        # Calculate subtotal
        subtotal = 0
        invoice_items = []

        for item_data in items:
            product = self.db.query(Product).filter(
                Product.id == item_data['product_id']
            ).first()

            if not product:
                raise ValueError(f"Product {item_data['product_id']} not found")

            quantity = item_data['quantity']
            item_subtotal = product.price * quantity

            invoice_item = InvoiceItem(
                product_id=product.id,
                product_name=product.name,
                product_price=product.price,
                quantity=quantity,
                unit=product.unit,
                subtotal=item_subtotal
            )
            invoice_items.append(invoice_item)
            subtotal += item_subtotal

        # Calculate total (prevent negative totals)
        total = max(0, subtotal - discount + tax)

        # Generate invoice number
        invoice_number = self.generate_invoice_number()

        # Calculate initial payment status
        paid_amount = 0
        if status == 'paid':
            paid_amount = total
        
        remaining_amount = max(0, total - paid_amount)

        # Create invoice
        invoice = Invoice(
            invoice_number=invoice_number,
            customer_id=customer_id,
            customer_name=customer_name,
            normalized_customer_name=normalize_vietnamese(customer_name) if customer_name else None,
            customer_phone=customer_phone,
            normalized_customer_phone=normalize_phone(customer_phone) if customer_phone else None,
            customer_address=customer_address,
            subtotal=subtotal,
            discount=discount,
            tax=tax,
            total=total,
            paid_amount=paid_amount,
            remaining_amount=remaining_amount,
            status=status,
            payment_method=payment_method,
            notes=notes
        )

        self.db.add(invoice)
        self.db.flush()  # Get invoice ID

        # Add items
        for item in invoice_items:
            item.invoice_id = invoice.id
            self.db.add(item)

        self.db.commit()
        self.db.refresh(invoice)

        return invoice

    def get_invoice(self, invoice_id: int) -> Optional[Invoice]:
        """Get invoice by ID."""
        return self.db.query(Invoice).filter(Invoice.id == invoice_id).first()

    def update_invoice(
        self,
        invoice_id: int,
        items: List[Dict],
        customer_id: Optional[int] = None,
        customer_name: Optional[str] = None,
        customer_phone: Optional[str] = None,
        customer_address: Optional[str] = None,
        discount: float = 0,
        tax: float = 0,
        payment_method: Optional[str] = None,
        notes: Optional[str] = None,
        status: Optional[str] = None
    ) -> Optional[Invoice]:
        """
        Update an existing invoice.
        
        Business rules:
        - Only invoices in 'pending' or 'processing' status can be edited.
        - Processing invoices can only move to 'pending' or 'paid' and require items when finalized.
        - Pending invoices cannot change status via this endpoint.

        Args:
            invoice_id: Invoice ID
            items: List of dicts with 'product_id' and 'quantity'
            customer_id: Customer ID (if registered customer)
            customer_name: Customer name (for one-time customers)
            customer_phone: Customer phone
            customer_address: Customer address
            discount: Discount amount
            tax: Tax amount
            payment_method: Payment method
            notes: Additional notes

        Returns:
            Updated Invoice object or None if not found
        """
        invoice = self.get_invoice(invoice_id)
        if not invoice:
            return None

        current_status = invoice.status
        target_status = status or current_status

        if target_status not in ALLOWED_STATUSES:
            raise ValueError("Trạng thái hóa đơn không hợp lệ.")

        # Prevent editing invoices with any payment (partial or full)
        if invoice.paid_amount > 0:
            raise ValueError(
                f"Không thể chỉnh sửa hóa đơn đã có thanh toán ({invoice.paid_amount:,.0f}đ). "
                "Vui lòng sử dụng tính năng 'Hoàn trả hóa đơn' nếu cần điều chỉnh."
            )

        # Prevent editing invoices that have been exported
        if invoice.exported_at is not None:
            export_date = invoice.exported_at.strftime('%d/%m/%Y %H:%M')
            raise ValueError(
                f"Không thể chỉnh sửa hóa đơn đã xuất file ({export_date}). "
                "Hóa đơn đã xuất là tài liệu chính thức không được phép sửa đổi."
            )

        if current_status in ['paid', 'cancelled']:
            raise ValueError("Chỉ có thể chỉnh sửa hóa đơn ở trạng thái chờ thanh toán hoặc chờ xử lý.")

        if current_status == 'pending' and target_status != 'pending':
            raise ValueError("Không thể đổi trạng thái khi chỉnh sửa hóa đơn chờ thanh toán.")

        if current_status == 'processing':
            if target_status not in ['processing', 'pending', 'paid']:
                raise ValueError("Hóa đơn chờ xử lý chỉ có thể hoàn tất về trạng thái chờ thanh toán hoặc đã thanh toán.")
            if target_status != 'processing' and len(items) == 0:
                raise ValueError("Cần ít nhất một sản phẩm để hoàn tất hóa đơn chờ xử lý.")

        if len(items) == 0:
            raise ValueError("Hóa đơn cần ít nhất một sản phẩm.")

        # Remove existing items and recalculate totals
        self.db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice_id).delete()
        self.db.flush()

        subtotal = 0
        for item_data in items:
            product = self.db.query(Product).filter(Product.id == item_data['product_id']).first()
            if not product:
                raise ValueError(f"Product {item_data['product_id']} not found")

            quantity = item_data['quantity']
            item_subtotal = product.price * quantity

            invoice_item = InvoiceItem(
                invoice_id=invoice_id,
                product_id=product.id,
                product_name=product.name,
                product_price=product.price,
                quantity=quantity,
                unit=product.unit,
                subtotal=item_subtotal
            )
            self.db.add(invoice_item)
            subtotal += item_subtotal

        total = subtotal - discount + tax

        # Recalculate remaining amount based on new total and existing paid amount
        invoice.remaining_amount = max(0, total - invoice.paid_amount)
        
        invoice.customer_id = customer_id
        invoice.customer_name = customer_name
        invoice.normalized_customer_name = normalize_vietnamese(customer_name) if customer_name else None
        invoice.customer_phone = customer_phone
        invoice.normalized_customer_phone= normalize_phone(customer_phone) if customer_phone else None
        invoice.customer_address = customer_address
        invoice.subtotal = subtotal
        invoice.discount = discount
        invoice.tax = tax
        invoice.total = total
        invoice.status = target_status
        invoice.payment_method = payment_method
        invoice.notes = notes
        invoice.updated_at = get_vn_time()

        self.db.commit()
        self.db.refresh(invoice)

        return invoice

    def get_invoice_by_number(self, invoice_number: str) -> Optional[Invoice]:
        """Get invoice by invoice number."""
        return self.db.query(Invoice).filter(
            Invoice.invoice_number == invoice_number
        ).first()

    def search_invoices(
        self,
        customer_id: Optional[int] = None,
        status: Optional[Union[str, List[str]]] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        invoice_number: Optional[str] = None,
        customer_name: Optional[str] = None,
        customer_phone: Optional[str] = None,
        limit: int = 50,
        offset: int = 0
    ) -> Tuple[List[Invoice], int, bool, Optional[int]]:
        """
        Search invoices with filters and Vietnamese normalization support.
        Uses joinedload to prevent N+1 queries.
        
        Multi-tier search strategy for text fields:
        1. Search original fields (invoice_number, customer_name, customer_phone)
        2. Search normalized fields (normalized_customer_name, normalized_customer_phone)
        
        This allows searching "nguyen" to find invoices with customer "Nguyễn Văn A",
        or "0123" to find various phone formats like "(012) 345-6789".
        
        Args:
            customer_id: Filter by customer ID
            status: Filter by invoice status (single or list)
            start_date: Filter by start date
            end_date: Filter by end date
            invoice_number: Search by invoice number (partial match)
            customer_name: Search by customer name (with Vietnamese normalization)
            customer_phone: Search by customer phone (with format normalization)
            limit: Page size
            offset: Offset for pagination
        """
        from typing import Union
        from sqlalchemy.orm import joinedload

        safe_limit, safe_offset = self._normalize_page_params(limit, offset)
        filters = []
        base_filters = []  # filters without search OR for fuzzy fallback

        if customer_id:
            filters.append(Invoice.customer_id == customer_id)
            base_filters.append(Invoice.customer_id == customer_id)

        if status:
            if isinstance(status, list):
                filters.append(Invoice.status.in_(status))
                base_filters.append(Invoice.status.in_(status))
            else:
                filters.append(Invoice.status == status)
                base_filters.append(Invoice.status == status)

        if start_date:
            filters.append(Invoice.created_at >= start_date)
            base_filters.append(Invoice.created_at >= start_date)

        if end_date:
            filters.append(Invoice.created_at <= end_date)
            base_filters.append(Invoice.created_at <= end_date)

        # Search filters - use OR logic for search fields
        search_filters = []
        
        if invoice_number:
            # Invoice number search (original field only)
            search_filters.append(Invoice.invoice_number.ilike(f"%{invoice_number}%"))

        if customer_name:
            # Customer name search with Vietnamese normalization
            normalized_name = normalize_vietnamese(customer_name)
            search_filters.extend([
                Invoice.customer_name.ilike(f"%{customer_name}%"),
                Invoice.normalized_customer_name.ilike(f"%{normalized_name}%")
            ])

        if customer_phone:
            # Customer phone search with format normalization
            normalized_phone_query = normalize_phone(customer_phone)
            search_filters.append(Invoice.customer_phone.ilike(f"%{customer_phone}%"))
            # Avoid adding a wildcard '%%' when query has no digits
            if normalized_phone_query:
                search_filters.append(Invoice.normalized_customer_phone.ilike(f"%{normalized_phone_query}%"))

        # Add OR search filters if any
        if search_filters:
            filters.append(or_(*search_filters))

        query = self.db.query(Invoice)

        # Eagerly load invoice items to prevent N+1 queries
        query = query.options(joinedload(Invoice.items))

        strict_query = query
        if filters:
            strict_query = strict_query.filter(and_(*filters))

        total = strict_query.count()
        strict_results = strict_query.order_by(Invoice.created_at.desc(), Invoice.id.desc()).offset(safe_offset).limit(safe_limit + 1).all()
        has_more = len(strict_results) > safe_limit
        items = strict_results[:safe_limit]

        # Fuzzy matching for typo tolerance if a search term is present and we still have room
        search_term_present = any([invoice_number, customer_name, customer_phone])
        normalized_name = normalize_vietnamese(customer_name) if customer_name else ""
        normalized_phone_query = normalize_phone(customer_phone) if customer_phone else ""
        search_term_lower = (invoice_number or customer_name or customer_phone or "").lower()

        seen_ids = {inv.id for inv in items}
        fuzzy_total = 0

        if search_term_present and len(items) < safe_limit:
            base_query = self.db.query(Invoice).options(joinedload(Invoice.items))
            if base_filters:
                base_query = base_query.filter(and_(*base_filters))

            candidates = base_query.order_by(Invoice.created_at.desc(), Invoice.id.desc()).limit(300).all()
            fuzzy_results: List[Tuple[Invoice, int]] = []

            for inv in candidates:
                if inv.id in seen_ids:
                    continue

                scores = []
                if invoice_number:
                    scores.append(fuzz.partial_ratio(invoice_number.lower(), inv.invoice_number.lower()))
                if customer_name and inv.customer_name:
                    scores.append(fuzz.partial_ratio(search_term_lower, inv.customer_name.lower()))
                if normalized_name and inv.normalized_customer_name:
                    scores.append(fuzz.partial_ratio(normalized_name, inv.normalized_customer_name))
                if normalized_phone_query and inv.normalized_customer_phone:
                    scores.append(fuzz.partial_ratio(normalized_phone_query, inv.normalized_customer_phone))

                best_score = max(scores) if scores else 0
                if best_score >= 70:
                    fuzzy_results.append((inv, best_score))
                    seen_ids.add(inv.id)

            fuzzy_results.sort(key=lambda item: item[1], reverse=True)
            fuzzy_matches = [inv for inv, _ in fuzzy_results]
            fuzzy_total = len(fuzzy_matches)

            remaining_slots = safe_limit - len(items)
            if remaining_slots > 0:
                items.extend(fuzzy_matches[:remaining_slots])

            if not has_more:
                has_more = fuzzy_total > remaining_slots

        total += fuzzy_total
        next_offset = safe_offset + safe_limit if has_more else None
        return items, total, has_more, next_offset

    def update_invoice_status(self, invoice_id: int, status: str) -> Optional[Invoice]:
        """
        Update invoice status.
        
        Business rule: Only invoices with 'pending' status can be updated.
        Once an invoice is marked as 'paid' or 'cancelled', it cannot be changed.
        Processing invoices must be finalized through update_invoice, not via direct status change.
        
        Args:
            invoice_id: Invoice ID
            status: New status (pending, paid, cancelled)
            
        Returns:
            Updated invoice or None if not found
            
        Raises:
            ValueError: If trying to update a finalized invoice (paid or cancelled)
        """
        invoice = self.get_invoice(invoice_id)
        if not invoice:
            return None

        if status not in ALLOWED_STATUSES:
            raise ValueError("Trạng thái hóa đơn không hợp lệ.")
        if status == 'processing':
            raise ValueError("Không thể chuyển trạng thái sang 'chờ xử lý' thông qua API.")
        
        # Business rule: Only pending invoices can have their status changed
        if invoice.status in ['paid', 'cancelled']:
            raise ValueError(
                f"Cannot update status of {invoice.status} invoice. "
                "Only pending invoices can be updated."
            )
        if invoice.status == 'processing':
            raise ValueError("Không thể cập nhật trạng thái của hóa đơn đang chờ xử lý. Vui lòng hoàn tất hóa đơn trước.")

        # NEW: Validate cancel with payments
        if status == 'cancelled' and invoice.paid_amount > 0:
            raise ValueError(
                f"Không thể hủy hóa đơn đã thanh toán {invoice.paid_amount:,.0f}đ. "
                f"Vui lòng sử dụng tính năng 'Hoàn trả hóa đơn' thay thế."
            )

        # If status is being changed to 'paid', we need to settle the debt
        if status == 'paid' and invoice.remaining_amount > 0.01:
            try:
                # If invoice has a customer, create a payment record
                if invoice.customer_id:
                    from services.payment_service import PaymentService
                    payment_service = PaymentService(self.db, self.output_dir)
                    
                    payment_service.record_payment(
                        customer_id=invoice.customer_id,
                        amount=invoice.remaining_amount,
                        payment_method=invoice.payment_method or 'cash',
                        invoice_ids=[invoice.id],
                        notes="Tự động thanh toán khi cập nhật trạng thái hóa đơn"
                    )
                    # record_payment will update the invoice amounts and status logic
                    # We reload the invoice to get updated values
                    self.db.refresh(invoice)
                    return invoice
                else:
                    # Walk-in customer (no customer_id): Just update amounts manually
                    # Since we can't create a Payment record without a customer_id
                    invoice.paid_amount = invoice.total
                    invoice.remaining_amount = 0
            except Exception as e:
                raise ValueError(f"Lỗi khi tự động thanh toán: {str(e)}")

        invoice.status = status
        invoice.updated_at = get_vn_time()

        self.db.commit()
        self.db.refresh(invoice)

        return invoice

    def cleanup_processing_invoices(self, max_age_hours: int = 24) -> int:
        """
        Delete processing invoices older than the given threshold.

        Args:
            max_age_hours: Age threshold in hours

        Returns:
            Number of invoices deleted
        """
        cutoff = get_vn_time() - timedelta(hours=max_age_hours)
        stale_invoices = self.db.query(Invoice).filter(
            Invoice.status == 'processing',
            Invoice.created_at <= cutoff
        ).all()

        deleted_count = len(stale_invoices)
        for invoice in stale_invoices:
            self.db.delete(invoice)

        if deleted_count > 0:
            self.db.commit()

        return deleted_count

    def generate_pdf(self, invoice_id: int, company_name: str = "Voi Store") -> str:
        """
        Generate PDF invoice.

        Args:
            invoice_id: Invoice ID
            company_name: Company name to display on invoice

        Returns:
            Path to generated PDF file
        """
        invoice = self.get_invoice(invoice_id)
        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Set exported_at on first export
        if invoice.exported_at is None:
            invoice.exported_at = get_vn_time()
            self.db.commit()
            self.db.refresh(invoice)

        # Create PDF filename
        pdf_filename = f"{invoice.invoice_number}.pdf"
        pdf_path = os.path.join(self.output_dir, pdf_filename)

        # Create PDF
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=FONT_BOLD,
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
            alignment=1  # Center
        )

        elements.append(Paragraph(f"<b>{company_name}</b>", title_style))
        elements.append(Paragraph(f"HÓA ĐƠN BÁN HÀNG", title_style))
        elements.append(Spacer(1, 12))

        # Invoice info
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=10
        )
        elements.append(Paragraph(f"<b>Số hóa đơn:</b> {invoice.invoice_number}", info_style))
        elements.append(Paragraph(f"<b>Ngày:</b> {invoice.created_at.strftime('%d/%m/%Y %H:%M')}", info_style))

        if invoice.customer_name:
            elements.append(Paragraph(f"<b>Khách hàng:</b> {invoice.customer_name}", info_style))
        if invoice.customer_phone:
            elements.append(Paragraph(f"<b>Số điện thoại:</b> {invoice.customer_phone}", info_style))
        if invoice.customer_address:
            elements.append(Paragraph(f"<b>Địa chỉ:</b> {invoice.customer_address}", info_style))

        elements.append(Spacer(1, 20))

        # Items table
        # Create header with Paragraphs for better text wrapping
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName=FONT_BOLD,
            fontSize=8,
            textColor=colors.whitesmoke,
            alignment=1  # Center
        )
        
        table_data = [[
            Paragraph('<b>STT</b>', header_style),
            Paragraph('<b>Sản phẩm</b>', header_style),
            Paragraph('<b>Đơn giá<br/>(VNĐ)</b>', header_style),
            Paragraph('<b>SL</b>', header_style),
            Paragraph('<b>Đơn vị</b>', header_style),
            Paragraph('<b>Thành tiền<br/>(VNĐ)</b>', header_style)
        ]]

        # Product name style for wrapping
        product_style = ParagraphStyle(
            'ProductStyle',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=8,
            alignment=0  # Left align
        )

        for idx, item in enumerate(invoice.items, 1):
            table_data.append([
                str(idx),
                Paragraph(item.product_name, product_style),
                f"{item.product_price:,.0f}",
                str(item.quantity),
                item.unit,
                f"{item.subtotal:,.0f}"
            ])
        
        # Calculate number of item rows for grid styling
        num_items = len(invoice.items)

        # Add totals
        # We place the label in the first column and will merge columns 0-4
        row_idx = num_items + 1
        footer_styles = []

        # Subtotal
        table_data.append(['Tạm tính:', '', '', '', '', f"{invoice.subtotal:,.0f}"])
        footer_styles.append(('SPAN', (0, row_idx), (4, row_idx)))
        footer_styles.append(('ALIGN', (0, row_idx), (4, row_idx), 'RIGHT'))
        row_idx += 1

        if invoice.discount > 0:
            table_data.append(['Giảm giá:', '', '', '', '', f"-{invoice.discount:,.0f}"])
            footer_styles.append(('SPAN', (0, row_idx), (4, row_idx)))
            footer_styles.append(('ALIGN', (0, row_idx), (4, row_idx), 'RIGHT'))
            row_idx += 1

        if invoice.tax > 0:
            table_data.append(['Thuế:', '', '', '', '', f"{invoice.tax:,.0f}"])
            footer_styles.append(('SPAN', (0, row_idx), (4, row_idx)))
            footer_styles.append(('ALIGN', (0, row_idx), (4, row_idx), 'RIGHT'))
            row_idx += 1
        
        # Total - span entire width with left and right alignment
        # Create a style for left-aligned total label
        total_label_style = ParagraphStyle(
            'TotalLabelStyle',
            parent=styles['Normal'],
            fontName=FONT_BOLD,
            fontSize=11,
            alignment=0  # Left align
        )
        total_value_style = ParagraphStyle(
            'TotalValueStyle',
            parent=styles['Normal'],
            fontName=FONT_BOLD,
            fontSize=11,
            alignment=2  # Right align
        )
        table_data.append(['Tổng cộng:', '', '', '', '', f"{invoice.total:,.0f}"])
        footer_styles.append(('SPAN', (1, row_idx), (4, row_idx)))  # Span columns 1-4 only
        footer_styles.append(('ALIGN', (0, row_idx), (0, row_idx), 'LEFT'))  # Left align label
        footer_styles.append(('ALIGN', (5, row_idx), (5, row_idx), 'RIGHT'))  # Right align value
        footer_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), FONT_BOLD))
        footer_styles.append(('FONTSIZE', (0, row_idx), (-1, row_idx), 11))
        # Add top border for Total row
        footer_styles.append(('LINEABOVE', (0, row_idx), (-1, row_idx), 1, colors.black))
        footer_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 15))
        footer_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 10))


        # Adjusted column widths for larger numbers:
        # STT: 25, Product: 135, Price(9 digits): 70, Qty(3 digits): 30, Unit: 55, Subtotal(12 digits): 135
        table = Table(table_data, colWidths=[25, 135, 70, 30, 55, 135])

        
        # Base styles
        table_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),  # Vertically center header
            ('VALIGN', (0, 1), (-1, num_items), 'TOP'),  # Top align product rows
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, num_items), colors.beige),
            ('GRID', (0, 0), (-1, num_items), 1, colors.black),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'), # Right align prices and totals
            ('ALIGN', (1, 1), (1, num_items), 'LEFT'), # Left align product names
            ('FONTNAME', (0, -1), (-1, -1), FONT_BOLD),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('FONTNAME', (0, 1), (-1, -2), FONT_NORMAL),
            ('FONTSIZE', (0, 1), (-1, num_items), 8),
        ]
        
        # Add footer styles
        table_styles.extend(footer_styles)
        
        table.setStyle(TableStyle(table_styles))

        elements.append(table)
        elements.append(Spacer(1, 15))
        
        # Amount in words - Create a nice box for better visibility
        amount_text = number_to_vietnamese_text(invoice.total)
        
        # Create styles for the amount in words box
        label_style = ParagraphStyle(
            'AmountLabel',
            parent=styles['Normal'],
            fontName=FONT_BOLD,
            fontSize=10,
            textColor=colors.HexColor('#2C3E50'),
            alignment=0
        )
        
        value_style = ParagraphStyle(
            'AmountValue',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=10,
            textColor=colors.HexColor('#16A085'),
            alignment=0,
            leading=14
        )
        
        # Create a table for the amount in words with nice styling
        amount_box_data = [[
            Paragraph('<b>Thành tiền bằng chữ:</b>', label_style),
            Paragraph(f'<i>{amount_text}</i>', value_style)
        ]]
        
        amount_box = Table(amount_box_data, colWidths=[120, 330])
        amount_box.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F8F5')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2C3E50')),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#16A085')),
            ('LINEAFTER', (0, 0), (0, 0), 1, colors.HexColor('#BDC3C7')),
        ]))
        
        elements.append(amount_box)
        elements.append(Spacer(1, 20))

        # Notes
        if invoice.notes:
            elements.append(Paragraph(f"<b>Ghi chú:</b> {invoice.notes}", info_style))
            elements.append(Spacer(1, 12))

        # Signature section
        elements.append(Spacer(1, 30))
        
        signature_style = ParagraphStyle(
            'SignatureStyle',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=11,
            alignment=1  # Center alignment
        )
        
        signature_data = [
            [
                Paragraph('<b>Người nhận hàng</b><br/>(Ký, ghi rõ họ tên)', signature_style),
                Paragraph('<b>Người bán hàng</b><br/>(Ký, ghi rõ họ tên)', signature_style)
            ],
            ['', '']  # Empty row for signature space
        ]
        
        signature_table = Table(signature_data, colWidths=[225, 225])
        signature_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'TOP'),
            ('VALIGN', (0, 1), (-1, 1), 'BOTTOM'),
            ('FONTNAME', (0, 0), (-1, 0), FONT_NORMAL),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 60),  # Space for signature
        ]))
        
        elements.append(signature_table)

        # Build PDF
        def draw_watermark(canvas, doc):
            """Draw watermark on PDF page."""
            canvas.saveState()
            project_root = Path(__file__).resolve().parents[2]
            logo_path = project_root / "react-frontend" / "public" / "Image_bqzqd5bqzqd5bqzq.png"
            if logo_path.exists():
                # Center of A4 page
                page_width, page_height = A4
                image_width = 100 * mm
                image_height = 100 * mm
                x = (page_width - image_width) / 2
                y = (page_height - image_height) / 2
                
                canvas.setFillAlpha(0.1) # Transparency
                canvas.drawImage(str(logo_path), x, y, width=image_width, height=image_height, mask='auto', preserveAspectRatio=True)
                
            canvas.restoreState()

        doc.build(elements, onFirstPage=draw_watermark, onLaterPages=draw_watermark)

        return pdf_path

    def generate_excel(self, invoice_id: int, company_name: str = "Voi Store") -> str:
        """
        Generate Excel invoice.

        Args:
            invoice_id: Invoice ID
            company_name: Company name

        Returns:
            Path to generated Excel file
        """
        invoice = self.get_invoice(invoice_id)
        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Set exported_at on first export
        if invoice.exported_at is None:
            invoice.exported_at = get_vn_time()
            self.db.commit()
            self.db.refresh(invoice)

        # Create Excel filename
        excel_filename = f"{invoice.invoice_number}.xlsx"
        excel_path = os.path.join(self.output_dir, excel_filename)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hóa đơn"

        # Styles
        title_font = Font(size=18, bold=True)
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = company_name
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = "HÓA ĐƠN BÁN HÀNG"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Invoice info
        row = 4
        ws[f'A{row}'] = f"Số hóa đơn: {invoice.invoice_number}"
        row += 1
        ws[f'A{row}'] = f"Ngày: {invoice.created_at.strftime('%d/%m/%Y %H:%M')}"
        row += 1
        if invoice.customer_name:
            ws[f'A{row}'] = f"Khách hàng: {invoice.customer_name}"
            row += 1
        if invoice.customer_phone:
            ws[f'A{row}'] = f"Số điện thoại: {invoice.customer_phone}"
            row += 1
        if invoice.customer_address:
            ws[f'A{row}'] = f"Địa chỉ: {invoice.customer_address}"
            row += 1

        row += 1

        # Headers
        headers = ['STT', 'Sản phẩm', 'Đơn giá', 'Số lượng', 'Đơn vị', 'Thành tiền']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Items
        for idx, item in enumerate(invoice.items, 1):
            row += 1
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=item.product_name).border = border
            ws.cell(row=row, column=3, value=item.product_price).border = border
            ws.cell(row=row, column=4, value=item.quantity).border = border
            ws.cell(row=row, column=5, value=item.unit).border = border
            ws.cell(row=row, column=6, value=item.subtotal).border = border

        # Totals
        row += 2
        ws.cell(row=row, column=5, value="Tạm tính:")
        ws.cell(row=row, column=6, value=invoice.subtotal)

        if invoice.discount > 0:
            row += 1
            ws.cell(row=row, column=5, value="Giảm giá:")
            ws.cell(row=row, column=6, value=-invoice.discount)

        if invoice.tax > 0:
            row += 1
            ws.cell(row=row, column=5, value="Thuế:")
            ws.cell(row=row, column=6, value=invoice.tax)

        row += 1
        ws.cell(row=row, column=5, value="Tổng cộng:").font = Font(bold=True)
        ws.cell(row=row, column=6, value=invoice.total).font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

        wb.save(excel_path)

        return excel_path

    def print_invoice(self, invoice_id: int, company_name: str = "Voi Store"):
        """
        Print invoice (generates PDF and opens it).

        Args:
            invoice_id: Invoice ID
            company_name: Company name
        """
        pdf_path = self.generate_pdf(invoice_id, company_name)

        # Open PDF with default viewer
        import platform
        import subprocess

        system = platform.system()
        if system == 'Windows':
            os.startfile(pdf_path)
        elif system == 'Darwin':  # macOS
            subprocess.run(['open', pdf_path])
        else:  # Linux
            subprocess.run(['xdg-open', pdf_path])

    def get_statistics(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict:
        """
        Get invoice statistics with debt tracking using SQL aggregation for performance.

        Revenue Definitions:
        - Total Revenue: Sum of all invoice totals (paid + pending, excluding cancelled)
        - Collected Amount: Sum of paid_amount across all active invoices
        - Outstanding Debt: Sum of remaining_amount across all active invoices
        - Customers with Debt: Count of distinct customers with remaining_amount > 0

        Args:
            start_date: Start date filter
            end_date: End date filter

        Returns:
            Dictionary with statistics
        """
        from sqlalchemy import func, case, distinct

        filters = []
        # Only include paid and pending invoices (exclude cancelled and processing)
        filters.append(Invoice.status.in_(['pending', 'paid']))
        if start_date:
            filters.append(Invoice.created_at >= start_date)
        if end_date:
            filters.append(Invoice.created_at <= end_date)

        query = self.db.query(Invoice)
        if filters:
            query = query.filter(and_(*filters))

        # Use SQL aggregation instead of Python loops for performance
        results = query.with_entities(
            func.count(Invoice.id).label('total_invoices'),

            # Total revenue = sum of all invoice totals (paid + pending)
            func.sum(Invoice.total).label('total_revenue'),

            # Collected amount = sum of paid_amount
            func.sum(Invoice.paid_amount).label('collected_amount'),

            # Outstanding debt = sum of remaining_amount
            func.sum(Invoice.remaining_amount).label('outstanding_debt'),

            # Legacy metrics (keep for backward compatibility)
            func.sum(case(
                (Invoice.status == 'paid', Invoice.total),
                else_=0
            )).label('paid_revenue'),
            func.sum(case(
                (Invoice.status == 'pending', Invoice.total),
                else_=0
            )).label('pending_revenue'),

            # Invoice counts by status
            func.count(case((Invoice.status == 'paid', 1))).label('paid_invoices'),
            func.count(case((Invoice.status == 'pending', 1))).label('pending_invoices'),
            func.count(case((Invoice.status == 'cancelled', 1))).label('cancelled_invoices'),
        ).first()

        # Query for customers with debt (separate query for clarity)
        customers_with_debt_query = self.db.query(
            func.count(distinct(Invoice.customer_id))
        ).filter(
            Invoice.status.in_(['pending', 'paid']),
            Invoice.remaining_amount > 0,
            Invoice.customer_id.isnot(None)  # Exclude walk-in customers
        )

        if start_date:
            customers_with_debt_query = customers_with_debt_query.filter(
                Invoice.created_at >= start_date
            )
        if end_date:
            customers_with_debt_query = customers_with_debt_query.filter(
                Invoice.created_at <= end_date
            )

        customers_with_debt = customers_with_debt_query.scalar() or 0

        # Extract and sanitize results
        total_invoices = results.total_invoices or 0
        total_revenue = float(results.total_revenue or 0)
        collected_amount = float(results.collected_amount or 0)
        outstanding_debt = float(results.outstanding_debt or 0)
        paid_invoices = results.paid_invoices or 0

        # Count invoices with debt
        invoices_with_debt = self.db.query(Invoice).filter(
            Invoice.status.in_(['pending', 'paid']),
            Invoice.remaining_amount > 0
        )
        if start_date:
            invoices_with_debt = invoices_with_debt.filter(Invoice.created_at >= start_date)
        if end_date:
            invoices_with_debt = invoices_with_debt.filter(Invoice.created_at <= end_date)
        invoices_with_debt_count = invoices_with_debt.count()

        return {
            'total_invoices': total_invoices,
            'paid_invoices': paid_invoices,
            'pending_invoices': results.pending_invoices or 0,
            'cancelled_invoices': results.cancelled_invoices or 0,

            # Total revenue now includes both paid + pending
            'total_revenue': total_revenue,

            # New: Breakdown of collected vs outstanding
            'collected_amount': collected_amount,
            'outstanding_debt': outstanding_debt,

            # Legacy fields (keep for backward compatibility)
            'pending_revenue': float(results.pending_revenue or 0),
            'average_order_value': total_revenue / total_invoices if total_invoices > 0 else 0,

            # Debt tracking
            'total_debt': outstanding_debt,  # Populate existing field
            'invoices_with_debt': invoices_with_debt_count,
            'customers_with_debt': customers_with_debt,
        }
