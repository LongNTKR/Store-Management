"""Debt report generation service for customer reconciliation."""

import os
from typing import List, Dict, Optional
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
import pytz

from database.models import Invoice, Customer, Payment, PaymentAllocation
from config import Config

# ReportLab imports for PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, Frame, PageTemplate
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Timezone
VN_TZ = pytz.timezone('Asia/Ho_Chi_Minh')

def get_vn_time():
    """Get current time in Vietnam timezone (UTC+7)."""
    return datetime.now(VN_TZ)

# Font setup for Vietnamese PDF
FONT_DIR = os.path.join(os.path.dirname(__file__), 'fonts')
FONT_NORMAL = 'DejaVuSans'
FONT_BOLD = 'DejaVuSans-Bold'

# Register fonts if available
try:
    normal_font_path = os.path.join(FONT_DIR, 'DejaVuSans.ttf')
    bold_font_path = os.path.join(FONT_DIR, 'DejaVuSans-Bold.ttf')

    if os.path.exists(normal_font_path):
        pdfmetrics.registerFont(TTFont(FONT_NORMAL, normal_font_path))
    else:
        FONT_NORMAL = 'Helvetica'

    if os.path.exists(bold_font_path):
        pdfmetrics.registerFont(TTFont(FONT_BOLD, bold_font_path))
    else:
        FONT_BOLD = 'Helvetica-Bold'
except Exception as e:
    # Fallback to Helvetica if DejaVu fonts not available
    FONT_NORMAL = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'
    print(f"Warning: Could not load DejaVu fonts, using Helvetica. Error: {e}")


class DebtReportService:
    """Service for generating debt reconciliation reports."""

    def __init__(self, db_session: Session, output_dir: str):
        """
        Initialize debt report service.

        Args:
            db_session: Database session
            output_dir: Directory to save reports
        """
        self.db = db_session
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def calculate_aging_buckets(
        self,
        invoices: List[Invoice],
        as_of_date: Optional[datetime] = None
    ) -> List[Dict]:
        """
        Calculate aging buckets for invoices.

        Args:
            invoices: List of invoices with remaining debt
            as_of_date: Date to calculate aging from (default: now)

        Returns:
            List of aging bucket dicts with structure:
            {
                'bucket_label': '0-30 ngày',
                'bucket_key': '0-30',
                'invoice_count': 2,
                'total_amount': 1500000.0,
                'invoices': [Invoice, ...]
            }
        """
        if as_of_date is None:
            as_of_date = get_vn_time()

        # Define buckets
        buckets = [
            {'key': '0-30', 'label': '0-30 ngày', 'min_days': 0, 'max_days': 30},
            {'key': '30-60', 'label': '30-60 ngày', 'min_days': 30, 'max_days': 60},
            {'key': '60-90', 'label': '60-90 ngày', 'min_days': 60, 'max_days': 90},
            {'key': '90+', 'label': 'Trên 90 ngày', 'min_days': 90, 'max_days': float('inf')}
        ]

        # Initialize bucket results
        bucket_results = []
        for bucket in buckets:
            bucket_results.append({
                'bucket_label': bucket['label'],
                'bucket_key': bucket['key'],
                'invoice_count': 0,
                'total_amount': 0.0,
                'invoices': []
            })

        # Classify invoices into buckets
        for invoice in invoices:
            if invoice.remaining_amount <= 0:
                continue

            # Calculate days old
            try:
                # Handle timezone-aware and naive datetime
                inv_date = invoice.created_at
                if inv_date.tzinfo is None and as_of_date.tzinfo is not None:
                    # Make invoice date timezone-aware
                    inv_date = VN_TZ.localize(inv_date)
                elif inv_date.tzinfo is not None and as_of_date.tzinfo is None:
                    # Make as_of_date timezone-aware
                    as_of_date = VN_TZ.localize(as_of_date)

                days_old = (as_of_date - inv_date).days
            except:
                # Fallback: compare naive datetimes
                inv_naive = invoice.created_at.replace(tzinfo=None) if invoice.created_at.tzinfo else invoice.created_at
                as_of_naive = as_of_date.replace(tzinfo=None) if as_of_date.tzinfo else as_of_date
                days_old = (as_of_naive - inv_naive).days

            # Find appropriate bucket
            for i, bucket in enumerate(buckets):
                if bucket['min_days'] <= days_old < bucket['max_days']:
                    bucket_results[i]['invoice_count'] += 1
                    bucket_results[i]['total_amount'] += invoice.remaining_amount
                    bucket_results[i]['invoices'].append(invoice)
                    break

        return bucket_results

    def generate_debt_pdf(
        self,
        customer_id: int,
        company_name: Optional[str] = None
    ) -> str:
        """
        Generate PDF debt reconciliation report for customer.

        Args:
            customer_id: Customer ID
            company_name: Company name (default from Config)

        Returns:
            Path to generated PDF file

        Raises:
            ValueError: If customer not found
        """
        # Get customer
        customer = self.db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise ValueError(f"Không tìm thấy khách hàng với ID {customer_id}")

        if company_name is None:
            company_name = Config.COMPANY_NAME

        # Get ALL invoices for customer (paid, unpaid, cancelled)
        invoices = self.db.query(Invoice).filter(
            Invoice.customer_id == customer_id
        ).order_by(Invoice.created_at.desc()).all()  # Most recent first

        # Calculate totals (now includes all invoices, not just unpaid)
        # Exclude cancelled invoices from debt calculations
        total_debt = sum(inv.remaining_amount for inv in invoices if inv.remaining_amount > 0 and inv.status != 'cancelled')
        total_invoices = len(invoices)
        total_unpaid = len([inv for inv in invoices if inv.remaining_amount > 0 and inv.status != 'cancelled'])

        # Create PDF filename
        now = get_vn_time()
        pdf_filename = f"Cong-no-{customer.name.replace(' ', '-')}-{now.strftime('%Y%m%d')}.pdf"
        pdf_path = os.path.join(self.output_dir, pdf_filename)

        # Create PDF
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=20*mm, leftMargin=15*mm, rightMargin=15*mm)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=FONT_BOLD,
            fontSize=18,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=20,
            alignment=TA_CENTER
        )

        elements.append(Paragraph(f"<b>{company_name}</b>", title_style))
        elements.append(Paragraph(f"BÁO CÁO ĐỐI CHIẾU CÔNG NỢ", title_style))
        elements.append(Spacer(1, 10))

        # Report date
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=10,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Ngày báo cáo: {now.strftime('%d/%m/%Y %H:%M')}", info_style))
        elements.append(Spacer(1, 15))

        # Customer info
        customer_style = ParagraphStyle(
            'CustomerStyle',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=11
        )
        elements.append(Paragraph(f"<b>Khách hàng:</b> {customer.name}", customer_style))
        if customer.phone:
            elements.append(Paragraph(f"<b>Số điện thoại:</b> {customer.phone}", customer_style))
        if customer.address:
            elements.append(Paragraph(f"<b>Địa chỉ:</b> {customer.address}", customer_style))
        elements.append(Spacer(1, 15))

        # Summary section
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Heading2'],
            fontName=FONT_BOLD,
            fontSize=14,
            textColor=colors.HexColor('#E74C3C'),
            spaceAfter=10
        )
        elements.append(Paragraph("<b>TỔNG QUAN CÔNG NỢ</b>", summary_style))

        summary_data = [
            ['Tổng số hóa đơn:', str(total_invoices)],
            ['Số hóa đơn chưa thanh toán đủ:', str(total_unpaid)],
            ['Tổng công nợ:', f"{total_debt:,.0f} VNĐ"]
        ]

        summary_table = Table(summary_data, colWidths=[250, 150], hAlign='LEFT')
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), FONT_NORMAL),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('FONTNAME', (0, -1), (-1, -1), FONT_BOLD),
            ('FONTSIZE', (0, -1), (-1, -1), 13),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#E74C3C')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Invoice list section
        elements.append(Paragraph("<b>CHI TIẾT HÓA ĐƠN</b>", summary_style))

        # Create styles for columns that need wrapping (Đã trả, Còn lại, Trạng thái, Ghi chú)
        cell_style_center = ParagraphStyle(
            'CellCenter',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=8,
            leading=10,
            alignment=TA_CENTER
        )

        cell_style_right = ParagraphStyle(
            'CellRight',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=8,
            leading=10,
            alignment=TA_RIGHT
        )

        cell_style_left = ParagraphStyle(
            'CellLeft',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=8,
            leading=10,
            alignment=TA_LEFT
        )

        # Status translation map
        status_map = {
            'pending': 'Chưa thanh toán',
            'paid': 'Đã thanh toán',
            'processing': 'Đang xử lý',
            'cancelled': 'Đã hủy'
        }

        # 8 columns now including Trạng thái
        invoice_table_data = [['STT', 'Số hóa đơn', 'Ngày', 'Tổng tiền', 'Đã trả', 'Còn lại', 'Trạng thái', 'Ghi chú']]

        for idx, invoice in enumerate(invoices, 1):
            # First 4 columns: Plain strings (NO wrapping - single line only)
            stt_str = str(idx)
            invoice_num_str = invoice.invoice_number or ''
            datetime_str = invoice.created_at.strftime('%d/%m/%Y %H:%M')
            total_str = f"{invoice.total:,.0f} VNĐ"

            # Last 4 columns: Paragraph objects (WITH wrapping capability)
            paid_para = Paragraph(f"{invoice.paid_amount:,.0f} VNĐ", cell_style_right)
            remaining_para = Paragraph(f"{invoice.remaining_amount:,.0f} VNĐ", cell_style_right)

            status_text = status_map.get(invoice.status, invoice.status or '')
            status_para = Paragraph(status_text, cell_style_center)

            notes_text = invoice.notes or ''
            notes_para = Paragraph(notes_text, cell_style_left)

            invoice_table_data.append([
                stt_str,           # STT - plain string
                invoice_num_str,   # Số hóa đơn - plain string
                datetime_str,      # Ngày - plain string
                total_str,         # Tổng tiền - plain string
                paid_para,         # Đã trả - Paragraph (wrappable)
                remaining_para,    # Còn lại - Paragraph (wrappable)
                status_para,       # Trạng thái - Paragraph (wrappable)
                notes_para         # Ghi chú - Paragraph (wrappable)
            ])

        # 8 columns, adjusted widths for A4 Landscape (~842pt width total, ~85pt margins L/R -> ~730pt usable)
        # Weights: STT:30, Số HĐ:90, Ngày:80, Tổng:80, Trả:80, Còn:80, Trạng thái:90, Ghi chú:200
        invoice_table = Table(invoice_table_data, colWidths=[30, 90, 80, 80, 80, 80, 90, 200])

        # Updated table style for mixed string/Paragraph cells
        invoice_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

            # Data rows - Font styling
            ('FONTNAME', (0, 1), (-1, -1), FONT_NORMAL),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),

            # Alignment for plain string columns (columns 0-3: STT, Số HĐ, Ngày, Tổng tiền)
            ('ALIGN', (0, 1), (2, -1), 'CENTER'),  # STT, Số HĐ, Ngày - center
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),   # Tổng tiền - right

            # Note: Columns 4-7 use Paragraph which handles alignment internally via styles

            # Vertical alignment for all cells
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),

            # Padding
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(invoice_table)
        elements.append(Spacer(1, 20))

        # Footer note
        elements.append(Spacer(1, 30))
        note_style = ParagraphStyle(
            'NoteStyle',
            parent=styles['Normal'],
            fontName=FONT_NORMAL,
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(
            f"Báo cáo được tạo tự động từ hệ thống {Config.APP_NAME}",
            note_style
        ))

        # Build PDF
        doc.build(elements)

        return pdf_path

    def generate_debt_excel(
        self,
        customer_id: int
    ) -> str:
        """
        Generate Excel debt reconciliation report for customer.

        Args:
            customer_id: Customer ID

        Returns:
            Path to generated Excel file

        Raises:
            ValueError: If customer not found
        """
        # Get customer
        customer = self.db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise ValueError(f"Không tìm thấy khách hàng với ID {customer_id}")

        # Get ALL invoices for customer (paid, unpaid, cancelled)
        invoices = self.db.query(Invoice).filter(
            Invoice.customer_id == customer_id
        ).order_by(Invoice.created_at.desc()).all()  # Most recent first

        # Create Excel file
        now = get_vn_time()
        excel_filename = f"Cong-no-{customer.name.replace(' ', '-')}-{now.strftime('%Y%m%d')}.xlsx"
        excel_path = os.path.join(self.output_dir, excel_filename)

        wb = Workbook()

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Tổng quan"

        ws_summary['A1'] = f"BÁO CÁO CÔNG NỢ - {customer.name.upper()}"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A2'] = f"Ngày báo cáo: {now.strftime('%d/%m/%Y %H:%M')}"

        ws_summary['A4'] = "Thông tin khách hàng:"
        ws_summary['A4'].font = Font(bold=True)
        ws_summary['A5'] = f"Tên: {customer.name}"
        ws_summary['A6'] = f"Số điện thoại: {customer.phone or 'N/A'}"
        ws_summary['A7'] = f"Địa chỉ: {customer.address or 'N/A'}"

        total_debt = sum(inv.remaining_amount for inv in invoices if inv.remaining_amount > 0)
        total_unpaid = len([inv for inv in invoices if inv.remaining_amount > 0])

        ws_summary['A9'] = "Tổng công nợ:"
        ws_summary['A9'].font = Font(bold=True, size=12)
        ws_summary['B9'] = total_debt
        ws_summary['B9'].number_format = '#,##0 "VNĐ"'
        ws_summary['B9'].font = Font(bold=True, size=12, color="E74C3C")

        ws_summary['A10'] = "Tổng số hóa đơn:"
        ws_summary['B10'] = len(invoices)

        ws_summary['A11'] = "Số hóa đơn có công nợ:"
        ws_summary['B11'] = total_unpaid

        # Sheet 2: Invoice Details (8 columns including Trạng thái)
        ws_invoices = wb.create_sheet("Chi tiết hóa đơn")

        # Status translation map
        status_map = {
            'pending': 'Chưa thanh toán',
            'paid': 'Đã thanh toán',
            'processing': 'Đang xử lý',
            'cancelled': 'Đã hủy'
        }

        headers = ['STT', 'Số hóa đơn', 'Ngày', 'Tổng tiền', 'Đã trả', 'Còn lại', 'Trạng thái', 'Ghi chú']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_invoices.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for idx, invoice in enumerate(invoices, 2):
            # Format datetime with time
            datetime_str = invoice.created_at.strftime('%d/%m/%Y %H:%M')
            status_text = status_map.get(invoice.status, invoice.status or '')

            ws_invoices.cell(row=idx, column=1, value=idx-1)  # STT
            ws_invoices.cell(row=idx, column=2, value=invoice.invoice_number)  # Số hóa đơn
            ws_invoices.cell(row=idx, column=3, value=datetime_str)  # Ngày
            ws_invoices.cell(row=idx, column=4, value=invoice.total)  # Tổng tiền
            ws_invoices.cell(row=idx, column=5, value=invoice.paid_amount)  # Đã trả
            ws_invoices.cell(row=idx, column=6, value=invoice.remaining_amount)  # Còn lại
            ws_invoices.cell(row=idx, column=7, value=status_text)  # Trạng thái
            ws_invoices.cell(row=idx, column=8, value=invoice.notes or '')  # Ghi chú

            # Format currency (columns 4, 5, 6)
            for col in [4, 5, 6]:
                ws_invoices.cell(row=idx, column=col).number_format = '#,##0'

            # Center align for STT, Số HĐ, Ngày, Trạng thái
            for col in [1, 2, 3, 7]:
                ws_invoices.cell(row=idx, column=col).alignment = Alignment(horizontal='center')

            # Right align for money columns
            for col in [4, 5, 6]:
                ws_invoices.cell(row=idx, column=col).alignment = Alignment(horizontal='right')

            # Left align for notes with wrap_text
            ws_invoices.cell(row=idx, column=8).alignment = Alignment(horizontal='left', wrap_text=True)

            # Add borders to all 8 columns
            for col in range(1, 9):
                ws_invoices.cell(row=idx, column=col).border = border

        # Auto-size columns
        for ws in [ws_summary, ws_invoices]:
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Save workbook
        wb.save(excel_path)

        return excel_path
